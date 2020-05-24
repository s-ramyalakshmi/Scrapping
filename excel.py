from openpyxl import load_workbook

wb = load_workbook('/home/rishikeshwar/Codes/Scrapping/excelsheet.xlsx')
ws = wb['Data']
for row in range(2, 3):
	for col in range(1, 10):
		ws.cell(row=row, column=col).value = 1

wb.save('document_template.xlsx')