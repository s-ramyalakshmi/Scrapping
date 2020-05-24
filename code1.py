import requests
import re
from bs4 import BeautifulSoup
import numpy as np
from openpyxl import load_workbook
from googlesearch import search

#https://www.webmd.com/lung/understanding-tuberculosis-basics#2
#https://www.webmd.com/a-to-z-guides/typhoid-fever#2
#https://www.webmd.com/search/search_results/default.aspx?query=dengue%20fever

diseases = []
symptoms = []

def getUrl(website, disease):
    return list(search(website + ' ' + disease + ' symptoms', stop=5))[0]

def get_diseases():
    wb = load_workbook('/home/rishikeshwar/Codes/Scrapping/excelsheet.xlsx')
    ws = wb['Symptoms']

    for row in range(2, 10000):
        k = ws.cell(row=row, column=1).value
        if k == None or k == '':
            break
        diseases.append(k)
        

def get_symptoms():
    wb = load_workbook('/home/rishikeshwar/Codes/Scrapping/excelsheet.xlsx')
    ws = wb['Symptoms']

    for row in range(2, 10000):
        k = ws.cell(row=row, column=2).value
        if k == None or k == '':
            break
        symptoms.append(k)


def trade_spider(stri):
    sendback = []
    url = getUrl('webmd', stri)
    print(url)
    source_code = requests.get(url)
    plain_text = source_code.text
    soup = BeautifulSoup(plain_text, 'lxml')
    for link in soup.findAll('div', {'class': 'active-page'}):
        for sec in link.findAll('section'):
            head = sec.find(re.compile('h[1-9]'))
            if head: #and bool(re.search(('(.)*(S|s)ymptoms(.)*'), head.string)):
                for li in sec.findAll(['li']):
                    text = li.text
                    sendback.append(text)
    return sendback

def check_symptom(getSymptoms):
    ans = []
    for i in symptoms:
        sympt = '(.)*' + i.lower().replace(' ', '(.)*') + '(.)*'
        goingToPush = 0
        for j in getSymptoms:
            if(bool(re.search(sympt, j.lower()))):
                goingToPush = 1
                break
        ans.append(goingToPush)
    return ans

if __name__ == '__main__':

    get_diseases()
    get_symptoms()

    wb = load_workbook('/home/rishikeshwar/Codes/Scrapping/excelsheet.xlsx')
    ws = wb['Data']
    col = 3

    for i in symptoms:
        ws.cell(row=1, column=col).value = i
        col += 1

    row = 2
    for i in diseases:
        getSymptoms = trade_spider(i)
        getans = check_symptom(getSymptoms)
        ws.cell(row=row, column=1).value = row - 1
        ws.cell(row=row, column=2).value = i
        for col in range(3, 3 + len(getans)):
            ws.cell(row=row, column=col).value = getans[col - 3]
        row += 1
        print('Disease ' + i + ' Done')


    wb.save('document_template.xlsx')